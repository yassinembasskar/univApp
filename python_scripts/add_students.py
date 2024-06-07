import sys
import json
from openpyxl import load_workbook

def process_excel(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    if sheet['A1'].value != 'email' or sheet['B1'].value != 'password':
        return json.dumps({"error": "A1 is not 'email' or A2 is not 'password'"})
    
    result = []

    row = 2
    while True:
        email_cell = sheet.cell(row=row, column=1)
        password_cell = sheet.cell(row=row, column=2)

        if not email_cell.value or not password_cell.value:
            break

        result.append({
            'email': email_cell.value,
            'password': password_cell.value
        })
        row += 1

    return json.dumps(result, indent=2)

if __name__ == "__main__":
    file_path = sys.argv[1]
    print(process_excel(file_path))